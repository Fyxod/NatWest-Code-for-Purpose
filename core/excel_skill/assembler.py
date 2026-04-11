"""
Excel Assembler — builds .xlsx files from structured plans and data.

All operations are deterministic (no LLM calls). Uses openpyxl for
workbook creation with formatting, formulas, and charts.
"""

import os
import re
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from core.llm.output_schemas.excel_skill_outputs import (
    ChartSpec,
    ExcelSkillPlan,
    SheetSpec,
)

# ─── Styling constants ────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style="thin", color="1F3864"),
    right=Side(style="thin", color="D9E2F3"),
)

DATA_FONT = Font(name="Calibri", size=10)
DATA_BORDER = Border(
    bottom=Side(style="hair", color="D9E2F3"),
    right=Side(style="hair", color="D9E2F3"),
)
ALT_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

SUMMARY_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUMMARY_LABEL_FONT = Font(name="Calibri", bold=True, size=11)
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11)

# Number format mapping
FORMAT_MAP = {
    "number": "#,##0",
    "currency": "$#,##0.00",
    "percentage": "0.00%",
    "date": "yyyy-mm-dd",
    "text": "@",
}

MAX_EXCEL_SHEET_NAME_LENGTH = 31


def assemble_excel(
    plan: ExcelSkillPlan,
    sheet_data: Dict[str, pd.DataFrame],
    output_path: str,
) -> str:
    """
    Assemble an Excel workbook from a plan and extracted data.

    Args:
        plan: The ExcelSkillPlan with sheet/column/chart specs.
        sheet_data: Dict mapping sheet_name → DataFrame.
        output_path: Where to write the .xlsx file.

    Returns:
        The output file path.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    # Remove default sheet — we'll create named ones
    if wb.active:
        wb.remove(wb.active)

    # Build summary sheet first if requested
    if plan.summary_sheet:
        _build_summary_sheet(wb, plan, sheet_data)

    # Build data sheets
    sheet_name_map: Dict[str, str] = {}
    for sheet_spec in plan.sheets:
        df = sheet_data.get(sheet_spec.sheet_name, pd.DataFrame())
        actual_sheet_name = _resolve_sheet_name(
            sheet_spec.sheet_name, set(wb.sheetnames)
        )
        sheet_name_map[sheet_spec.sheet_name] = actual_sheet_name
        _build_data_sheet(wb, sheet_spec, df, actual_sheet_name)

    # Add charts
    if plan.charts:
        for chart_spec in plan.charts:
            _add_chart(wb, chart_spec, sheet_data, sheet_name_map)

    # If workbook has no sheets (edge case), add a placeholder
    if not wb.sheetnames:
        ws = wb.active or wb.create_sheet("Sheet1")
        ws["A1"] = "No data available"

    wb.save(output_path)
    return output_path


def _build_summary_sheet(
    wb: Workbook,
    plan: ExcelSkillPlan,
    sheet_data: Dict[str, pd.DataFrame],
) -> None:
    """Build a summary/dashboard sheet with key metrics."""
    ws = wb.create_sheet("Summary", 0)

    # Title
    ws["A1"] = plan.description or plan.file_name
    ws["A1"].font = SUMMARY_TITLE_FONT
    ws.merge_cells("A1:D1")

    row = 3
    for sheet_spec in plan.sheets:
        df = sheet_data.get(sheet_spec.sheet_name, pd.DataFrame())
        if df.empty:
            continue

        ws.cell(row=row, column=1, value=sheet_spec.sheet_name).font = Font(
            bold=True, size=12, color="2F5496"
        )
        row += 1

        # Row count
        ws.cell(row=row, column=1, value="Total Rows").font = SUMMARY_LABEL_FONT
        ws.cell(row=row, column=2, value=len(df)).font = SUMMARY_VALUE_FONT
        row += 1

        # Numeric column stats
        numeric_cols = df.select_dtypes(include=["number"]).columns
        for col in numeric_cols[:5]:  # limit to 5 metrics per sheet
            ws.cell(row=row, column=1, value=f"{col} (sum)").font = SUMMARY_LABEL_FONT
            ws.cell(row=row, column=2, value=df[col].sum()).font = SUMMARY_VALUE_FONT
            ws.cell(row=row, column=2).number_format = "#,##0.00"
            row += 1

            ws.cell(row=row, column=1, value=f"{col} (avg)").font = SUMMARY_LABEL_FONT
            ws.cell(row=row, column=2, value=df[col].mean()).font = SUMMARY_VALUE_FONT
            ws.cell(row=row, column=2).number_format = "#,##0.00"
            row += 1

        row += 1  # blank row between sheet summaries

    # Auto-width for summary
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20


def _build_data_sheet(
    wb: Workbook,
    sheet_spec: SheetSpec,
    df: pd.DataFrame,
    actual_sheet_name: str,
) -> None:
    """Build a single data sheet with formatting and formulas."""
    ws = wb.create_sheet(actual_sheet_name)

    if df.empty:
        ws["A1"] = f"No data available for: {sheet_spec.description}"
        return

    # Apply group_by aggregation if specified
    if sheet_spec.group_by and sheet_spec.aggregations:
        df = _apply_aggregation(df, sheet_spec.group_by, sheet_spec.aggregations)

    # Apply sorting
    if sheet_spec.sort_by:
        ascending = True
        sort_col = sheet_spec.sort_by
        if sort_col.startswith("-"):
            ascending = False
            sort_col = sort_col[1:]
        if sort_col in df.columns:
            df = df.sort_values(by=sort_col, ascending=ascending).reset_index(drop=True)

    # Determine column order from spec (if columns match)
    spec_col_names = [c.name for c in sheet_spec.columns if c.source == "sql"]
    if spec_col_names and all(c in df.columns for c in spec_col_names):
        # Use spec order, keeping extra columns at the end
        extra_cols = [c for c in df.columns if c not in spec_col_names]
        df = df[spec_col_names + extra_cols]

    # Write header row
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER

    # Write data rows
    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_clean_value(value))
            cell.font = DATA_FONT
            cell.border = DATA_BORDER

            # Alternating row fill
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Apply formula columns
    _apply_formula_columns(ws, sheet_spec, len(df))

    # Apply number formats from column specs
    _apply_number_formats(ws, sheet_spec, df)

    # Auto-filter
    if len(df) > 0:
        last_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"

    # Auto-width columns
    _auto_width(ws, df)

    # Freeze top row
    ws.freeze_panes = "A2"


def _apply_aggregation(
    df: pd.DataFrame,
    group_by: List[str],
    aggregations: dict,
) -> pd.DataFrame:
    """Apply pandas groupby aggregation (pivot-table style)."""
    try:
        # Filter to valid columns
        valid_group = [c for c in group_by if c in df.columns]
        valid_aggs = {k: v for k, v in aggregations.items() if k in df.columns}

        if not valid_group or not valid_aggs:
            return df

        result = df.groupby(valid_group, as_index=False).agg(valid_aggs)
        return result
    except Exception as e:
        print(f"[ExcelSkill:assembler] Aggregation error: {e}")
        return df


def _apply_formula_columns(
    ws,
    sheet_spec: SheetSpec,
    data_row_count: int,
) -> None:
    """Insert formula columns defined in the sheet spec."""
    existing_col_count = ws.max_column or 0

    for col_spec in sheet_spec.columns:
        if not col_spec.source.startswith("formula:"):
            continue

        formula_template = col_spec.source[len("formula:"):]

        # Find or create the column
        col_idx = None
        for c in range(1, existing_col_count + 1):
            if ws.cell(row=1, column=c).value == col_spec.name:
                col_idx = c
                break

        if col_idx is None:
            # Add new column
            existing_col_count += 1
            col_idx = existing_col_count
            cell = ws.cell(row=1, column=col_idx, value=col_spec.name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT

        # Apply formula to each data row
        for row in range(2, data_row_count + 2):
            # Replace row references in formula (e.g., B2 → B{row})
            formula = _adjust_formula_row(formula_template, row)
            ws.cell(row=row, column=col_idx, value=formula)


def _adjust_formula_row(formula_template: str, target_row: int) -> str:
    """
    Adjust cell references in a formula template to the target row.

    E.g., '=C2*D2' with target_row=5 becomes '=C5*D5'
    """
    # Match cell references like A2, B2, AA2 (only row 2 references)
    def replace_ref(match):
        col_ref = match.group(1)
        return f"{col_ref}{target_row}"

    return re.sub(r"([A-Z]+)2\b", replace_ref, formula_template)


def _apply_number_formats(
    ws,
    sheet_spec: SheetSpec,
    df: pd.DataFrame,
) -> None:
    """Apply number formats to columns based on their data_type or number_format spec."""
    col_map = {c.name: c for c in sheet_spec.columns}

    for col_idx, col_name in enumerate(df.columns, 1):
        spec = col_map.get(col_name)
        if not spec:
            continue

        # Custom format takes priority
        fmt = spec.number_format or FORMAT_MAP.get(spec.data_type)
        if not fmt or fmt == "@":
            continue

        for row in range(2, len(df) + 2):
            ws.cell(row=row, column=col_idx).number_format = fmt


def _auto_width(ws, df: pd.DataFrame) -> None:
    """Auto-size column widths based on content."""
    for col_idx, col_name in enumerate(df.columns, 1):
        # Start with header width
        max_width = len(str(col_name)) + 2

        # Sample first 50 rows for width estimation
        for row_idx in range(2, min(len(df) + 2, 52)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_width = max(max_width, min(len(str(cell_value)), 50))

        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max_width + 2


def _add_chart(
    wb: Workbook,
    chart_spec: ChartSpec,
    sheet_data: Dict[str, pd.DataFrame],
    sheet_name_map: Dict[str, str],
) -> None:
    """Add a chart to the specified sheet."""
    worksheet_name = sheet_name_map.get(chart_spec.sheet_name)
    if not worksheet_name:
        truncated_name = chart_spec.sheet_name[:MAX_EXCEL_SHEET_NAME_LENGTH]
        if truncated_name in wb.sheetnames:
            worksheet_name = truncated_name
    if not worksheet_name or worksheet_name not in wb.sheetnames:
        return

    ws = wb[worksheet_name]
    df = sheet_data.get(chart_spec.sheet_name, pd.DataFrame())

    if df.empty:
        return

    # Find column indices
    col_names = list(df.columns)
    x_idx = None
    y_indices = []

    if chart_spec.x_column in col_names:
        x_idx = col_names.index(chart_spec.x_column) + 1

    for y_col in chart_spec.y_columns:
        if y_col in col_names:
            y_indices.append(col_names.index(y_col) + 1)

    if x_idx is None or not y_indices:
        print(f"[ExcelSkill:assembler] Chart column not found: {chart_spec}")
        return

    data_rows = len(df) + 1  # +1 for header

    # Create chart
    chart_map = {
        "bar": BarChart,
        "column": BarChart,
        "line": LineChart,
        "pie": PieChart,
        "scatter": ScatterChart,
    }
    chart_class = chart_map.get(chart_spec.chart_type.lower(), BarChart)
    chart = chart_class()
    chart.title = chart_spec.title
    chart.style = 10
    chart.width = 20
    chart.height = 12

    # Categories (X axis)
    cats = Reference(ws, min_col=x_idx, min_row=2, max_row=data_rows)

    # Data series (Y axis)
    for y_idx in y_indices:
        data = Reference(ws, min_col=y_idx, min_row=1, max_row=data_rows)
        chart.add_data(data, titles_from_data=True)

    if not isinstance(chart, PieChart):
        chart.set_categories(cats)

    # Place chart below the data
    chart_anchor = f"{get_column_letter(len(col_names) + 2)}2"
    ws.add_chart(chart, chart_anchor)


def _clean_value(value):
    """Clean a cell value for Excel compatibility."""
    if pd.isna(value):
        return None
    if isinstance(value, str):
        # Strip unicode artifacts
        value = value.strip()
        if not value or value.lower() == "nan":
            return None
    return value


def _resolve_sheet_name(requested_name: str, existing_names: set[str]) -> str:
    """Normalize and deduplicate worksheet names within Excel's limits."""
    safe_name = re.sub(r'[\\/*?:\[\]]', "_", requested_name or "").strip() or "Sheet"
    base_name = safe_name[:MAX_EXCEL_SHEET_NAME_LENGTH]
    candidate = base_name
    suffix = 2

    while candidate in existing_names:
        suffix_text = f"_{suffix}"
        trimmed_base = base_name[: MAX_EXCEL_SHEET_NAME_LENGTH - len(suffix_text)]
        candidate = f"{trimmed_base}{suffix_text}"
        suffix += 1

    return candidate
